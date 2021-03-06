import json

import openpyxl


def readFromJson(file):
    with open(file, 'r', encoding='UTF-8') as fr:
        jsonData = json.load(fr)
    return jsonData


def writeToExcel(file):
    json = readFromJson(file)
    excel = openpyxl.Workbook()
    sheet1 = excel.create_sheet('sheet1', index=0)
    sheet2 = excel.create_sheet('sheet2', index=0)
    length = len(json)
    i = 0
    while i < length:
        eachLine = json[i]
        questions = eachLine['questions']
        answer = eachLine['answer']
        questionSize = len(questions)
        j = 0
        while j < questionSize:
            ques = questions[j]
            eachQues = ques['question']
            sheet1.cell(row=i + 1, column=j + 1, value=eachQues)
            if j == 0:
                sheet2.cell(row=i + 1, column=1, value=eachQues)
            j = j + 1
        sheet2.cell(row=i + 1, column=2, value=answer)
        i = i + 1
    excel.save('data2.xlsx')


if __name__ == '__main__':
    writeToExcel('C:/Users/YAE/Desktop/python_all/Python_Openpyxl/temp2')
